__author__ = 'Janne Parkkila'

print("Hello World!")

from openpyxl import load_workbook
import os

def LoadWorkbook(filename):
    filename = "\\" + filename
    try:
        workbook = load_workbook(os.getcwd() + filename)
        print("Loaded the workbook from" + os.getcwd() + filename)
        return workbook
    except FileNotFoundError:
        print("File " + filename + " was not found...")
        exit(0)


def ShowWorkbookStats(workbook):
    print("Workbook contains the following sheets:")
    print(workbook.get_sheet_names())


wb = LoadWorkbook("test.xlsx")
ShowWorkbookStats(wb)